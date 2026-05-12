from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook, Workbook

from app.data_extraction.contract_extract import ContractExtraction
from app.data_extraction.populate_master_workbook import (
    assert_no_bank_details_in_workbook,
    populate_master_workbook,
)


def _dummy_extraction() -> ContractExtraction:
    return ContractExtraction(
        staff_number="123456",
        title="Ms",
        full_name="Demo Lecturer",
        surname="Lecturer",
        highest_qualification="Master of Dummy Systems",
        id_or_passport_number="DUMMY-ID-123456",
        paye_number="DUMMY-PAYE-123456",
        physical_address="P.O. Box 000, Windhoek",
        contact_number="0810000000",
        campus="Windhoek Main Campus",
        contract_start_date="2026-02-01",
        contract_end_date="2026-06-05",
        course_codes=["CUS411S", "ICT521S"],
        course_names=["Computer User Skills", "Information Competence"],
        faculty="Computing and Informatics",
        department="Informatics",
        tariff_per_hour="460",
        active="Yes",
        source_file="dummy.pdf",
    )


def test_populate_master_workbook_creates_lecturer_and_course_rows(monkeypatch, tmp_path):
    contracts_dir = tmp_path / "contracts"
    contracts_dir.mkdir()
    (contracts_dir / "dummy.pdf").write_bytes(b"%PDF dummy")
    output = tmp_path / "real_master_data_draft.xlsx"

    def fake_extract(_path: Path):
        return _dummy_extraction(), tmp_path / "dummy.json"

    monkeypatch.setattr("app.data_extraction.populate_master_workbook.extract_contract_file", fake_extract)
    result = populate_master_workbook(contracts_dir, output)
    workbook = load_workbook(output)

    assert result["contracts_scanned"] == 1
    assert result["lecturers_extracted"] == 1
    assert workbook["Lecturers"]["A2"].value == "123456"
    assert workbook["Lecturers"]["C2"].value == "Demo Lecturer"
    assert workbook["Courses"]["A2"].value == "CUS411S"
    assert workbook["Courses"]["A3"].value == "ICT521S"


def test_populate_master_workbook_leaves_unavailable_sheets_blank(monkeypatch, tmp_path):
    contracts_dir = tmp_path / "contracts"
    contracts_dir.mkdir()
    (contracts_dir / "dummy.pdf").write_bytes(b"%PDF dummy")
    output = tmp_path / "real_master_data_draft.xlsx"

    def fake_extract(_path: Path):
        return _dummy_extraction(), tmp_path / "dummy.json"

    monkeypatch.setattr("app.data_extraction.populate_master_workbook.extract_contract_file", fake_extract)
    populate_master_workbook(contracts_dir, output)
    workbook = load_workbook(output)

    for sheet_name in ("Groups", "Students", "Group_Enrolments", "Timetable"):
        assert workbook[sheet_name].max_row == 1


def test_assert_no_bank_details_in_workbook_fails_on_bank_text(tmp_path):
    output = tmp_path / "bad.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lecturers"
    sheet["A1"] = "bank account number"
    workbook.save(output)

    with pytest.raises(ValueError, match="Bank detail text detected"):
        assert_no_bank_details_in_workbook(output)


def test_assert_no_bank_details_in_workbook_passes_for_valid_workbook(monkeypatch, tmp_path):
    contracts_dir = tmp_path / "contracts"
    contracts_dir.mkdir()
    (contracts_dir / "dummy.pdf").write_bytes(b"%PDF dummy")
    output = tmp_path / "valid.xlsx"

    def fake_extract(_path: Path):
        return _dummy_extraction(), tmp_path / "dummy.json"

    monkeypatch.setattr("app.data_extraction.populate_master_workbook.extract_contract_file", fake_extract)
    populate_master_workbook(contracts_dir, output)

    assert_no_bank_details_in_workbook(output)
