from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from app.data_extraction.contract_review_workbook import convert_review_to_master_workbook
from app.data_extraction.ocr_contract_extract import REVIEW_COLUMNS
from app.data_extraction.populate_master_workbook import assert_no_bank_details_in_workbook


def _review_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contract_Review"
    sheet.append(REVIEW_COLUMNS)
    reviewed = {
        "source_file": "reviewed.pdf",
        "extraction_status": "OCR draft extracted",
        "review_status": "Reviewed",
        "staff_number": "654321",
        "title": "Dr",
        "full_name": "Dummy Reviewer",
        "surname": "Reviewer",
        "highest_qualification": "Doctor of Dummy Analytics",
        "id_or_passport_number": "DUMMY-ID-654321",
        "paye_number": "DUMMY-PAYE-654321",
        "physical_address": "P.O. Box 000",
        "contact_number": "0811111111",
        "campus": "Windhoek Main Campus",
        "contract_start_date": "2026-02-01",
        "contract_end_date": "2026-06-05",
        "course_codes": "CUS411S, ICT521S",
        "course_names": "Computer User Skills, Information Competence",
        "faculty": "Computing and Informatics",
        "department": "Informatics",
        "tariff_per_hour": "460",
        "active": "Yes",
        "review_notes": "",
    }
    pending = reviewed | {"source_file": "pending.pdf", "review_status": "Needs review", "staff_number": "999999"}
    sheet.append([reviewed.get(column, "") for column in REVIEW_COLUMNS])
    sheet.append([pending.get(column, "") for column in REVIEW_COLUMNS])
    workbook.save(path)


def test_convert_review_to_master_only_includes_reviewed_or_approved_rows(tmp_path):
    review = tmp_path / "ocr_contract_review.xlsx"
    output = tmp_path / "real_master_data_from_review.xlsx"
    _review_workbook(review)

    result = convert_review_to_master_workbook(review, output)
    workbook = load_workbook(output)

    assert result["approved_rows"] == 1
    assert workbook["Lecturers"]["A2"].value == "654321"
    assert workbook["Lecturers"].max_row == 2
    assert workbook["Courses"]["A2"].value == "CUS411S"
    assert workbook["Courses"]["A3"].value == "ICT521S"


def test_convert_review_to_master_leaves_non_contract_sheets_blank(tmp_path):
    review = tmp_path / "ocr_contract_review.xlsx"
    output = tmp_path / "real_master_data_from_review.xlsx"
    _review_workbook(review)

    convert_review_to_master_workbook(review, output)
    workbook = load_workbook(output)

    for sheet_name in ("Groups", "Students", "Group_Enrolments", "Timetable"):
        assert workbook[sheet_name].max_row == 1


def test_convert_review_to_master_has_no_bank_detail_text(tmp_path):
    review = tmp_path / "ocr_contract_review.xlsx"
    output = tmp_path / "real_master_data_from_review.xlsx"
    _review_workbook(review)
    convert_review_to_master_workbook(review, output)

    assert_no_bank_details_in_workbook(output)


def test_review_converter_requires_contract_review_sheet(tmp_path):
    review = tmp_path / "bad_review.xlsx"
    Workbook().save(review)

    with pytest.raises(ValueError, match="Contract_Review"):
        convert_review_to_master_workbook(review, tmp_path / "out.xlsx")
