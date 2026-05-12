from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import TEMPLATES_DIR


TEMPLATE_PATH = TEMPLATES_DIR / "master_data_template.xlsx"

SHEET_COLUMNS = {
    "Lecturers": [
        "staff_number",
        "title",
        "full_name",
        "highest_qualification",
        "id_or_passport_number",
        "paye_number",
        "physical_address",
        "contact_number",
        "tariff_per_hour",
        "campus",
        "contract_start_date",
        "contract_end_date",
        "active",
    ],
    "Courses": [
        "course_code",
        "course_name",
        "faculty",
        "department",
        "budget_allocation",
        "active",
    ],
    "Groups": [
        "group_name",
        "course_code",
        "campus",
        "study_mode",
        "active",
    ],
    "Students": [
        "student_number",
        "surname",
        "initials",
        "full_name",
        "active",
    ],
    "Group_Enrolments": [
        "student_number",
        "group_name",
        "course_code",
        "active",
    ],
    "Timetable": [
        "staff_number",
        "group_name",
        "course_code",
        "day_of_week",
        "start_time",
        "end_time",
        "effective_start_date",
        "effective_end_date",
        "active",
    ],
    "Academic_Calendar": [
        "title",
        "start_date",
        "end_date",
        "calendar_type",
        "action",
        "allow_override",
    ],
}

META_SHEETS = ("Instructions", "Data_Dictionary", "Allowed_Values")
DATA_SHEETS = tuple(SHEET_COLUMNS)


SAMPLE_ROWS = {
    "Lecturers": [
        {
            "staff_number": "DUMMY-LECT-0001",
            "title": "Ms",
            "full_name": "Demo Template Lecturer",
            "highest_qualification": "Dummy Qualification",
            "id_or_passport_number": "DUMMY-ID-TEMPLATE-0001",
            "paye_number": "DUMMY-PAYE-TEMPLATE-0001",
            "physical_address": "P.O. Box 000, Windhoek",
            "contact_number": "0810000000",
            "tariff_per_hour": 410,
            "campus": "Windhoek Main Campus",
            "contract_start_date": "2026-02-01",
            "contract_end_date": "2026-02-28",
            "active": "yes",
        }
    ],
    "Courses": [
        {
            "course_code": "CUS411S",
            "course_name": "Computer User Skills",
            "faculty": "Faculty of Computing and Informatics",
            "department": "Department of Informatics and Journalism",
            "budget_allocation": "0183-0102",
            "active": "yes",
        }
    ],
    "Groups": [
        {
            "group_name": "Group 1",
            "course_code": "CUS411S",
            "campus": "Eenhana Satellite Campus",
            "study_mode": "Part-time",
            "active": "yes",
        },
        {
            "group_name": "Group 2",
            "course_code": "CUS411S",
            "campus": "Eenhana Satellite Campus",
            "study_mode": "Part-time",
            "active": "yes",
        },
    ],
    "Students": [
        {
            "student_number": "DUMMY-STU-0001",
            "surname": "Amadhila",
            "initials": "T.",
            "full_name": "T. Amadhila",
            "active": "yes",
        },
        {
            "student_number": "DUMMY-STU-0002",
            "surname": "Iita",
            "initials": "M.",
            "full_name": "M. Iita",
            "active": "yes",
        },
    ],
    "Group_Enrolments": [
        {
            "student_number": "DUMMY-STU-0001",
            "group_name": "Group 1",
            "course_code": "CUS411S",
            "active": "yes",
        },
        {
            "student_number": "DUMMY-STU-0002",
            "group_name": "Group 2",
            "course_code": "CUS411S",
            "active": "yes",
        },
    ],
    "Timetable": [
        {
            "staff_number": "DUMMY-LECT-0001",
            "group_name": "Group 1",
            "course_code": "CUS411S",
            "day_of_week": "Monday",
            "start_time": "08:00",
            "end_time": "10:00",
            "effective_start_date": "2026-02-01",
            "effective_end_date": "2026-02-28",
            "active": "yes",
        },
        {
            "staff_number": "DUMMY-LECT-0001",
            "group_name": "Group 2",
            "course_code": "CUS411S",
            "day_of_week": "Tuesday",
            "start_time": "10:00",
            "end_time": "12:00",
            "effective_start_date": "2026-02-01",
            "effective_end_date": "2026-02-28",
            "active": "yes",
        },
    ],
    "Academic_Calendar": [
        {
            "title": "Sample public holiday",
            "start_date": "2026-02-09",
            "end_date": "2026-02-09",
            "calendar_type": "public_holiday",
            "action": "exclude",
            "allow_override": "no",
        }
    ],
}

FIELD_DESCRIPTIONS = {
    "staff_number": ("Yes", "Lecturer personnel/staff number used as the stable lecturer key.", "200001", "Use the exact same value in Timetable."),
    "title": ("Yes", "Lecturer title.", "Ms", ""),
    "full_name": ("Yes", "Lecturer full name for registers and claim forms.", "Demo Lecturer", "Use dummy names in templates/tests."),
    "highest_qualification": ("Yes", "Highest qualification for the claim form.", "Dummy Qualification", ""),
    "id_or_passport_number": ("Yes", "ID or passport field required by claim form.", "DUMMY-ID-0001", "Do not put real values in sample templates."),
    "paye_number": ("Yes", "PAYE field required by claim form.", "DUMMY-PAYE-0001", "Do not put real values in sample templates."),
    "physical_address": ("Yes", "Physical/postal address field required by claim form.", "P.O. Box 000, Windhoek", "Use real values only in controlled real-use workbooks."),
    "contact_number": ("Yes", "Contact number field required by claim form.", "0810000000", "Template values are dummy placeholders."),
    "tariff_per_hour": ("Yes", "Hourly tariff used to calculate claim amount.", "410", "Must be greater than zero."),
    "campus": ("Yes", "Campus for lecturer/group/register output.", "Windhoek Main Campus", ""),
    "contract_start_date": ("Yes", "First claimable teaching date.", "2026-02-01", "YYYY-MM-DD."),
    "contract_end_date": ("Yes", "Last claimable teaching date.", "2026-02-28", "YYYY-MM-DD."),
    "active": ("Yes", "Whether the record should be used.", "Yes", "Allowed values are listed in Allowed_Values."),
    "course_code": ("Yes", "Stable course code used to link groups, timetable, and enrolments.", "CUS411S", "Use exact values consistently."),
    "course_name": ("Yes", "Course name for claim/register output.", "Computer User Skills", ""),
    "faculty": ("Yes", "Faculty name.", "Computing and Informatics", ""),
    "department": ("Yes", "Department name.", "Informatics and Journalism", ""),
    "budget_allocation": ("Yes", "Budget allocation code for claim form.", "0183-0102", ""),
    "group_name": ("Yes", "Teaching group name.", "Demo Group A", "Use exact values consistently."),
    "study_mode": ("Yes", "Mode of study.", "Part-time", "Allowed values are listed in Allowed_Values."),
    "student_number": ("Yes", "Student number used as stable student key.", "900000001", ""),
    "surname": ("Conditional", "Student surname.", "Amunyela", "Surname or full_name is required."),
    "initials": ("No", "Student initials.", "A.", ""),
    "day_of_week": ("Yes", "Teaching weekday.", "Monday", "Allowed values are listed in Allowed_Values."),
    "start_time": ("Yes", "Session start time.", "08:00", "HH:MM."),
    "end_time": ("Yes", "Session end time.", "09:00", "HH:MM and after start_time."),
    "effective_start_date": ("Yes", "First date timetable entry applies.", "2026-02-01", "YYYY-MM-DD."),
    "effective_end_date": ("Yes", "Last date timetable entry applies.", "2026-02-28", "YYYY-MM-DD."),
    "title_calendar": ("Yes", "Calendar entry title.", "Institutional closure", ""),
    "start_date": ("Yes", "Calendar period start date.", "2026-02-10", "YYYY-MM-DD."),
    "end_date": ("Yes", "Calendar period end date.", "2026-02-10", "YYYY-MM-DD."),
    "calendar_type": ("Yes", "Calendar category.", "institutional_closure", "Allowed values are listed in Allowed_Values."),
    "action": ("Yes", "Whether the period includes or excludes sessions.", "exclude", "Allowed values: include, exclude."),
    "allow_override": ("Yes", "Whether override is allowed.", "No", "Allowed values are listed in Allowed_Values."),
}


def _instructions_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["Purpose", "Use this workbook to prepare lecturer, course, group, student, enrolment, timetable, and academic calendar data for pt_claims_system."],
            ["Safety", "Do not use real bank details. This system must not store bank details."],
            ["Required fields", "Fill all required fields before import."],
            ["Dates", "Dates must use YYYY-MM-DD format."],
            ["Times", "Times must use HH:MM format."],
            ["Course links", "Use exact course_code values when linking groups, timetable, and enrolments."],
            ["Group links", "Use exact group_name values consistently across sheets."],
            ["After real import", "Do not run dev_reset or seed_data after importing real data."],
            ["Workflow 1", "Generate blank template."],
            ["Workflow 2", "Fill lecturer, course, group, student, enrolment, timetable, and calendar sheets."],
            ["Workflow 3", "Run dry-run validation."],
            ["Workflow 4", "Back up the current database."],
            ["Workflow 5", "Import the workbook."],
            ["Workflow 6", "Inspect imported data."],
            ["Workflow 7", "Generate sessions."],
            ["Workflow 8", "Generate documents."],
        ],
        columns=["topic", "instruction"],
    )


def _data_dictionary_df() -> pd.DataFrame:
    rows = []
    for sheet_name, columns in SHEET_COLUMNS.items():
        for column in columns:
            key = "title_calendar" if sheet_name == "Academic_Calendar" and column == "title" else column
            required, description, example, notes = FIELD_DESCRIPTIONS.get(
                key,
                ("No", f"{column} value for {sheet_name}.", "", ""),
            )
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "column_name": column,
                    "required": required,
                    "description": description,
                    "example": example,
                    "notes": notes,
                }
            )
    return pd.DataFrame(rows)


def _allowed_values_df() -> pd.DataFrame:
    allowed = {
        "active": ["Yes", "No", "TRUE", "FALSE", "1", "0"],
        "study_mode": ["Full-time", "Part-time", "Extra-curricular"],
        "day_of_week": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
        "calendar_type": ["public_holiday", "recess", "institutional_closure", "special_event"],
        "action": ["include", "exclude"],
        "allow_override": ["Yes", "No", "TRUE", "FALSE", "1", "0"],
    }
    return pd.DataFrame(
        [{"field": field, "allowed_value": value} for field, values in allowed.items() for value in values]
    )


def _autosize_columns(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def generate_master_data_template(output_path: str | Path = TEMPLATE_PATH) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _instructions_df().to_excel(writer, sheet_name="Instructions", index=False)
        _data_dictionary_df().to_excel(writer, sheet_name="Data_Dictionary", index=False)
        _allowed_values_df().to_excel(writer, sheet_name="Allowed_Values", index=False)
        for sheet_name in META_SHEETS:
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
            _autosize_columns(writer, sheet_name)

        for sheet_name, columns in SHEET_COLUMNS.items():
            pd.DataFrame(SAMPLE_ROWS[sheet_name], columns=columns).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            _autosize_columns(writer, sheet_name)
    return output


def main() -> None:
    path = generate_master_data_template()
    print(f"Generated master data template at {path}")


if __name__ == "__main__":
    main()
