from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from app.data_extraction.populate_master_workbook import (
    KNOWN_COURSES,
    _append_dict,
    _clear_data_rows,
    _scrub_workbook_bank_labels,
    assert_no_bank_details_in_workbook,
)
from app.master_data_template import SHEET_COLUMNS, generate_master_data_template


APPROVED_STATUSES = {"reviewed", "approved"}


def _sheet_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    if "Contract_Review" not in workbook.sheetnames:
        raise ValueError("Review workbook must contain Contract_Review sheet.")
    sheet = workbook["Contract_Review"]
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({str(header): value for header, value in zip(headers, row)})
    return rows


def _split_codes(value) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).replace(";", ",").split(",") if part.strip()]


def convert_review_to_master_workbook(review_file: str | Path, output: str | Path) -> dict:
    review_path = Path(review_file)
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _sheet_rows(review_path)
    approved_rows = [
        row for row in rows if str(row.get("review_status") or "").strip().lower() in APPROVED_STATUSES
    ]

    generate_master_data_template(output_path)
    workbook = load_workbook(output_path)
    _scrub_workbook_bank_labels(workbook)
    for sheet_name in ("Lecturers", "Courses", "Groups", "Students", "Group_Enrolments", "Timetable", "Academic_Calendar"):
        _clear_data_rows(workbook[sheet_name])

    detected_codes = set()
    for row in approved_rows:
        lecturer = {column: row.get(column, "") or "" for column in SHEET_COLUMNS["Lecturers"]}
        _append_dict(workbook["Lecturers"], SHEET_COLUMNS["Lecturers"], lecturer)
        detected_codes.update(_split_codes(row.get("course_codes")))

    for code in sorted(detected_codes):
        if code in KNOWN_COURSES:
            _append_dict(workbook["Courses"], SHEET_COLUMNS["Courses"], KNOWN_COURSES[code])

    workbook.save(output_path)
    assert_no_bank_details_in_workbook(output_path)
    return {
        "output_path": output_path,
        "review_rows": len(rows),
        "approved_rows": len(approved_rows),
        "course_rows": len([code for code in detected_codes if code in KNOWN_COURSES]),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert reviewed OCR contract workbook to master data workbook.")
    parser.add_argument("--review-file", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    result = convert_review_to_master_workbook(args.review_file, args.output)
    print(f"Output workbook path: {result['output_path']}")
    print(f"Review rows read: {result['review_rows']}")
    print(f"Approved/reviewed lecturer rows included: {result['approved_rows']}")
    print(f"Course rows included: {result['course_rows']}")
    print("Group, student, enrolment, timetable, and academic calendar sheets were left blank.")


if __name__ == "__main__":
    main()
