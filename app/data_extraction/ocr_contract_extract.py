from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.docx_utils import safe_filename_text
from app.data_extraction.contract_extract import ContractExtraction, _strip_bank_lines, parse_contract_text


REVIEW_COLUMNS = [
    "source_file",
    "extraction_status",
    "review_status",
    "staff_number",
    "title",
    "full_name",
    "surname",
    "highest_qualification",
    "id_or_passport_number",
    "paye_number",
    "physical_address",
    "contact_number",
    "campus",
    "contract_start_date",
    "contract_end_date",
    "course_codes",
    "course_names",
    "faculty",
    "department",
    "tariff_per_hour",
    "active",
    "review_notes",
]


def ocr_dependencies_available() -> tuple[bool, str]:
    try:
        import pytesseract  # noqa: F401
        import pdf2image  # noqa: F401
    except ImportError as exc:
        return False, f"Missing Python OCR dependency: {exc.name}"
    return True, "OCR Python dependencies available"


def parse_pages_range(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"\s*(\d+)(?:\s*-\s*(\d+))?\s*", value)
    if not match:
        raise ValueError("Pages must be formatted like 1-2 or 1")
    start = int(match.group(1))
    end = int(match.group(2) or start)
    if start < 1 or end < start:
        raise ValueError("Invalid page range")
    return start, end


def parse_ocr_contract_text(text: str, source_file: str = "") -> ContractExtraction:
    extraction = parse_contract_text(text, source_file=source_file)
    extraction.active = "Yes" if extraction.staff_number or extraction.full_name else ""
    return extraction


def review_notes_for_extraction(extraction: ContractExtraction, text: str) -> str:
    notes = ["Scanned OCR result requires manual review."]
    if extraction.staff_number and re.search(r"Registration Number[^\n]*if applicable", text, re.IGNORECASE):
        notes.append("Staff number cleaned from noisy OCR label.")
    if not extraction.staff_number:
        notes.append("No reliable staff number found.")
    if not extraction.title:
        notes.append("Title uncertain.")
    if extraction.bank_details_detected:
        notes.append("Bank details detected and ignored.")
    if not extraction.full_name and (extraction.surname or extraction.staff_number):
        notes.append("Full name uncertain; manual review required.")
    return " ".join(dict.fromkeys(notes))


def extract_text_with_ocr(pdf_path: Path, pages: tuple[int, int]) -> tuple[str, str]:
    available, message = ocr_dependencies_available()
    if not available:
        return "", message
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(str(pdf_path), first_page=pages[0], last_page=pages[1])
        text_parts = [pytesseract.image_to_string(image) for image in images]
        return "\n".join(text_parts), "OCR completed"
    except Exception as exc:
        return "", f"OCR failed: {exc}"


def _row_from_extraction(extraction: ContractExtraction, status: str, notes: str = "") -> list:
    review_notes = []
    if notes:
        review_notes.append(notes)
    if extraction.bank_details_detected:
        review_notes.append("Bank details detected and ignored.")
    return [
        Path(extraction.source_file).name,
        status,
        "Needs review",
        extraction.staff_number,
        extraction.title,
        extraction.full_name,
        extraction.surname,
        extraction.highest_qualification,
        extraction.id_or_passport_number,
        extraction.paye_number,
        extraction.physical_address,
        extraction.contact_number,
        extraction.campus,
        extraction.contract_start_date,
        extraction.contract_end_date,
        ", ".join(extraction.course_codes),
        ", ".join(extraction.course_names),
        extraction.faculty,
        extraction.department,
        extraction.tariff_per_hour,
        extraction.active,
        " ".join(dict.fromkeys(review_notes)),
    ]


def _format_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for column_cells in sheet.columns:
            width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 12), 48)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def create_review_workbook(
    contracts_dir: str | Path,
    output: str | Path,
    pages: tuple[int, int] = (1, 2),
    all_pages: bool = False,
    save_ocr_text: bool = False,
) -> dict:
    source_dir = Path(contracts_dir)
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["topic", "instruction"])
    instructions.append(["Purpose", "OCR-assisted contract review workbook for drafting lecturer/course data only."])
    instructions.append(["Review", "OCR is draft only. Review every row manually before creating a master import workbook."])
    instructions.append(["Safety", "Payment account details are intentionally excluded and must not be imported."])
    instructions.append(["Scope", "Groups, students, enrolments, and timetables are not populated from contracts."])
    review = workbook.create_sheet("Contract_Review")
    review.append(REVIEW_COLUMNS)

    pdf_files = sorted(source_dir.glob("*.pdf")) if source_dir.exists() else []
    unavailable_messages = set()
    bank_detected = False
    ocr_text_dir = output_path.parent / "extracted_contracts" / "ocr_text"
    for pdf_file in pdf_files:
        page_range = (1, 9999) if all_pages else pages
        text, status_message = extract_text_with_ocr(pdf_file, page_range)
        if not text.strip():
            unavailable_messages.add(status_message)
            extraction = ContractExtraction(source_file=str(pdf_file), machine_readable_text_found=False, active="")
            review.append(_row_from_extraction(extraction, "Manual review required", status_message))
            continue
        sanitized_text, text_bank_detected = _strip_bank_lines(text)
        if save_ocr_text:
            ocr_text_dir.mkdir(parents=True, exist_ok=True)
            (ocr_text_dir / f"{safe_filename_text(pdf_file.stem)}.txt").write_text(sanitized_text, encoding="utf-8")
        extraction = parse_ocr_contract_text(text, source_file=str(pdf_file))
        bank_detected = bank_detected or extraction.bank_details_detected or text_bank_detected
        review.append(_row_from_extraction(extraction, "OCR draft extracted", review_notes_for_extraction(extraction, text)))

    required_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    required_headers = {"staff_number", "full_name", "title", "tariff_per_hour", "contract_start_date", "contract_end_date"}
    header_lookup = {cell.value: cell.column for cell in review[1]}
    for row in range(2, review.max_row + 1):
        for header in required_headers:
            cell = review.cell(row=row, column=header_lookup[header])
            if cell.value in (None, ""):
                cell.fill = required_fill

    _format_workbook(workbook)
    workbook.save(output_path)
    return {
        "contracts_scanned": len(pdf_files),
        "output_path": output_path,
        "ocr_available": not unavailable_messages,
        "ocr_messages": sorted(unavailable_messages),
        "bank_details_detected": bank_detected,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Create OCR-assisted contract review workbook.")
    parser.add_argument("--contracts-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--pages", default="1-2")
    parser.add_argument("--all-pages", action="store_true")
    parser.add_argument("--save-ocr-text", action="store_true")
    args = parser.parse_args()
    pages = parse_pages_range(args.pages)
    if args.all_pages:
        print("Warning: --all-pages is slower and may expose more irrelevant text.")
    if args.save_ocr_text:
        print("Raw OCR text may contain sensitive information. Use only for local debugging.")
    result = create_review_workbook(
        args.contracts_dir,
        args.output,
        pages=pages,
        all_pages=args.all_pages,
        save_ocr_text=args.save_ocr_text,
    )
    if not result["ocr_available"]:
        print("OCR dependencies not available. Created manual review workbook with blank fields.")
    if result["bank_details_detected"]:
        print("Bank details detected and ignored.")
    print(f"Contract files scanned: {result['contracts_scanned']}")
    print(f"Review workbook path: {result['output_path']}")


if __name__ == "__main__":
    main()
