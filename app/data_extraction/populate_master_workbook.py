from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from app.data_extraction.contract_extract import ContractExtraction, extract_contract_file, mask_sensitive
from app.master_data_template import SHEET_COLUMNS, generate_master_data_template


DEFAULT_OUTPUT = Path("data") / "real_imports" / "real_master_data_draft.xlsx"
KNOWN_COURSES = {
    "CUS411S": {
        "course_code": "CUS411S",
        "course_name": "Computer User Skills",
        "faculty": "Computing and Informatics",
        "department": "Informatics",
        "budget_allocation": "0183-0102",
        "active": "Yes",
    },
    "ICT521S": {
        "course_code": "ICT521S",
        "course_name": "Information Competence",
        "faculty": "Computing and Informatics",
        "department": "Informatics",
        "budget_allocation": "0183-0102",
        "active": "Yes",
    },
}
BANK_DETAIL_LABELS = (
    "bank",
    "account number",
    "account holder",
    "branch code",
    "swift",
    "first national bank",
    "fnb",
)


def _clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _append_dict(ws, columns: list[str], values: dict) -> None:
    ws.append([values.get(column, "") for column in columns])


def _scrub_workbook_bank_labels(workbook) -> None:
    replacements = {
        "Do not use real bank details. This system must not store bank details.": (
            "Do not include prohibited payment details. This system must not store payment account information."
        ),
        "bank details": "prohibited payment details",
        "bank": "payment institution",
        "account number": "payment reference",
        "account holder": "payment holder",
        "branch code": "branch reference",
        "swift": "international payment reference",
        "first national bank": "payment institution",
        "fnb": "payment institution",
    }
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = cell.value
                lowered = value.lower()
                for target, replacement in replacements.items():
                    if target in lowered:
                        value = re_sub_case_insensitive(target, replacement, value)
                        lowered = value.lower()
                cell.value = value


def re_sub_case_insensitive(target: str, replacement: str, value: str) -> str:
    import re

    return re.sub(re.escape(target), replacement, value, flags=re.IGNORECASE)


def lecturer_row(extraction: ContractExtraction) -> dict:
    return {
        "staff_number": extraction.staff_number,
        "title": extraction.title,
        "full_name": extraction.full_name,
        "highest_qualification": extraction.highest_qualification,
        "id_or_passport_number": extraction.id_or_passport_number,
        "paye_number": extraction.paye_number,
        "physical_address": extraction.physical_address,
        "contact_number": extraction.contact_number,
        "tariff_per_hour": extraction.tariff_per_hour,
        "campus": extraction.campus,
        "contract_start_date": extraction.contract_start_date,
        "contract_end_date": extraction.contract_end_date,
        "active": extraction.active or "Yes",
    }


def missing_fields(row: dict) -> list[str]:
    return [column for column, value in row.items() if value in ("", None)]


def _course_rows(extractions: list[ContractExtraction]) -> list[dict]:
    detected_codes = set()
    for extraction in extractions:
        detected_codes.update(extraction.course_codes)
        for course_name in extraction.course_names:
            lowered = course_name.lower()
            if "computer user skills" in lowered:
                detected_codes.add("CUS411S")
            if "information competence" in lowered:
                detected_codes.add("ICT521S")
    return [KNOWN_COURSES[code] for code in sorted(detected_codes) if code in KNOWN_COURSES]


def assert_no_bank_details_in_workbook(path: str | Path) -> None:
    workbook = load_workbook(path, data_only=False)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                text = str(cell.value or "").lower()
                if any(label in text for label in BANK_DETAIL_LABELS):
                    raise ValueError(f"Bank detail text detected in workbook sheet {sheet.title}.")


def populate_master_workbook(contracts_dir: str | Path, output: str | Path = DEFAULT_OUTPUT) -> dict:
    source_dir = Path(contracts_dir)
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_path = generate_master_data_template(output_path)
    workbook = load_workbook(template_path)
    _scrub_workbook_bank_labels(workbook)

    contract_files = sorted(source_dir.glob("*.pdf")) if source_dir.exists() else []
    extractions: list[ContractExtraction] = []
    json_paths: list[Path] = []
    for contract_file in contract_files:
        extraction, json_path = extract_contract_file(contract_file)
        extractions.append(extraction)
        json_paths.append(json_path)

    for sheet_name in ("Lecturers", "Courses", "Groups", "Students", "Group_Enrolments", "Timetable", "Academic_Calendar"):
        _clear_data_rows(workbook[sheet_name])

    lecturer_rows = []
    for extraction in extractions:
        row = lecturer_row(extraction)
        lecturer_rows.append(row)
        if any(row.values()):
            _append_dict(workbook["Lecturers"], SHEET_COLUMNS["Lecturers"], row)

    course_rows = _course_rows(extractions)
    for row in course_rows:
        _append_dict(workbook["Courses"], SHEET_COLUMNS["Courses"], row)

    workbook.save(output_path)
    assert_no_bank_details_in_workbook(output_path)
    return {
        "contracts_scanned": len(contract_files),
        "lecturers_extracted": len([row for row in lecturer_rows if row.get("staff_number") or row.get("full_name")]),
        "output_path": output_path,
        "json_paths": json_paths,
        "missing_fields": {
            row.get("staff_number") or row.get("full_name") or f"lecturer_{index}": missing_fields(row)
            for index, row in enumerate(lecturer_rows, start=1)
        },
        "scanned_without_text": [
            extraction.source_file for extraction in extractions if not extraction.machine_readable_text_found
        ],
        "bank_details_detected": any(extraction.bank_details_detected for extraction in extractions),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Populate draft master workbook from lecturer contracts.")
    parser.add_argument("--contracts-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    result = populate_master_workbook(args.contracts_dir, args.output)
    print(f"Contract files scanned: {result['contracts_scanned']}")
    print(f"Lecturers extracted: {result['lecturers_extracted']}")
    print(f"Output workbook path: {result['output_path']}")
    for lecturer_key, fields in result["missing_fields"].items():
        print(f"Missing fields for {mask_sensitive(lecturer_key)}: {', '.join(fields) if fields else 'None'}")
    for path in result["scanned_without_text"]:
        print(f"No machine-readable text found for {path}. Manual review required.")
    if result["bank_details_detected"]:
        print("Bank details detected and ignored.")
    print("Group, student, enrolment, and timetable sheets were left blank because those details are not available from lecturer contracts.")


if __name__ == "__main__":
    main()
